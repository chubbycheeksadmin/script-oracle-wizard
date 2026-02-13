#!/usr/bin/env python3
"""
Production data extractor v3 - Uses pdftotext to avoid iCloud file locks
"""

import json
import subprocess
import sys
from pathlib import Path
from typing import Dict, Any, Optional
import re
import warnings
warnings.filterwarnings('ignore')

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not available, Excel parsing disabled")

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False
    print("Warning: xlrd not available, old .xls parsing disabled")

def extract_text_from_pdf(pdf_path: Path, max_pages: int = 10) -> str:
    """Extract text from PDF using pdftotext command"""
    try:
        result = subprocess.run(
            ['pdftotext', '-l', str(max_pages), str(pdf_path), '-'],
            capture_output=True,
            text=True,
            timeout=30
        )
        return result.stdout[:50000]  # Limit to 50k chars
    except Exception as e:
        return f"[PDF extraction failed: {e}]"

def extract_features_from_script(text: str) -> Dict[str, Any]:
    """Extract key features from script text"""
    text_lower = text.lower()
    
    # Technique detection
    techniques = []
    technique_keywords = {
        'moco': ['moco', 'motion control', 'mo-co'],
        'drone': ['drone', 'aerial', 'uav'],
        'tracking': ['tracking shot', 'tracking', 'dolly track'],
        'vfx': ['vfx', 'visual effects', 'green screen', 'greenscreen', 'cgi'],
        'night_shoot': ['night shoot', 'night exterior', 'night int', 'night ext'],
        'underwater': ['underwater', 'submerged'],
        'crane': ['crane shot', 'crane', 'jib'],
        'steadicam': ['steadicam', 'steadi'],
        'handheld': ['handheld', 'hand held'],
        'slow_motion': ['slow motion', 'slow-motion', 'high speed', 'phantom'],
        'time_lapse': ['time lapse', 'time-lapse', 'timelapse']
    }
    
    for tech, keywords in technique_keywords.items():
        if any(kw in text_lower for kw in keywords):
            techniques.append(tech)
    
    # Location detection
    locations = []
    if 'studio' in text_lower or 'sound stage' in text_lower:
        locations.append('studio')
    if any(word in text_lower for word in ['exterior', 'ext.', 'outdoor', 'location']):
        locations.append('outdoor')
    if 'interior' in text_lower or 'int.' in text_lower:
        locations.append('indoor')
    
    # Shot/scene count
    shot_patterns = [
        r'shot\s+\d+',
        r'scene\s+\d+',
        r'sc\.\s*\d+',
    ]
    
    all_shots = []
    for pattern in shot_patterns:
        matches = re.findall(pattern, text_lower)
        all_shots.extend(matches)
    
    estimated_shots = len(set(all_shots)) if all_shots else None
    
    # Talent detection
    has_children = any(word in text_lower for word in ['child', 'kid', 'baby', 'infant'])
    has_animals = any(word in text_lower for word in ['dog', 'cat', 'horse', 'animal'])
    has_vehicles = any(word in text_lower for word in ['car', 'vehicle', 'truck', 'motorcycle'])
    
    return {
        'techniques': techniques,
        'locations': locations,
        'estimated_shots': estimated_shots,
        'has_children': has_children,
        'has_animals': has_animals,
        'has_vehicles': has_vehicles,
        'text_length': len(text)
    }

def extract_from_xlsx_budget(budget_path: Path) -> Dict[str, Any]:
    """Extract budget from .xlsx file"""
    if not HAS_OPENPYXL:
        return {'total_gbp': None, 'error': 'openpyxl not installed'}
    
    try:
        print(f"    Reading .xlsx: {budget_path.name}...")
        wb = openpyxl.load_workbook(budget_path, read_only=True, data_only=True)
        sheet = wb.active
        
        amounts = []
        for row in sheet.iter_rows(max_row=200, values_only=True):
            if not row:
                continue
            for cell in row:
                if cell is None:
                    continue
                
                cell_str = str(cell)
                
                # GBP pattern
                gbp_matches = re.findall(r'£\s*([\d,]+(?:\.\d{2})?)', cell_str)
                for m in gbp_matches:
                    try:
                        amount = float(m.replace(',', ''))
                        if 1000 < amount < 10000000:
                            amounts.append(amount)
                    except:
                        pass
                
                # Numeric cell
                if isinstance(cell, (int, float)) and 10000 < cell < 10000000:
                    amounts.append(cell)
        
        wb.close()
        
        if amounts:
            return {
                'total_gbp': round(max(amounts), 2),
                'source': 'xlsx',
                'amounts_found': len(amounts)
            }
        
        return {'total_gbp': None, 'note': 'No amounts in xlsx'}
        
    except Exception as e:
        return {'total_gbp': None, 'error': f'xlsx: {str(e)[:100]}'}

def extract_from_xls_budget(budget_path: Path) -> Dict[str, Any]:
    """Extract budget from old .xls file"""
    if not HAS_XLRD:
        return {'total_gbp': None, 'error': 'xlrd not installed'}
    
    try:
        print(f"    Reading .xls: {budget_path.name}...")
        wb = xlrd.open_workbook(budget_path)
        sheet = wb.sheet_by_index(0)
        
        amounts = []
        for row_idx in range(min(200, sheet.nrows)):
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                
                # Check text cells for GBP
                if cell.ctype == xlrd.XL_CELL_TEXT:
                    text = str(cell.value)
                    gbp_matches = re.findall(r'£\s*([\d,]+(?:\.\d{2})?)', text)
                    for m in gbp_matches:
                        try:
                            amount = float(m.replace(',', ''))
                            if 1000 < amount < 10000000:
                                amounts.append(amount)
                        except:
                            pass
                
                # Check numeric cells
                if cell.ctype == xlrd.XL_CELL_NUMBER:
                    if 10000 < cell.value < 10000000:
                        amounts.append(cell.value)
        
        if amounts:
            return {
                'total_gbp': round(max(amounts), 2),
                'source': 'xls',
                'amounts_found': len(amounts)
            }
        
        return {'total_gbp': None, 'note': 'No amounts in xls'}
        
    except Exception as e:
        return {'total_gbp': None, 'error': f'xls: {str(e)[:100]}'}

def extract_from_pdf_budget(budget_path: Path) -> Dict[str, Any]:
    """Extract budget from PDF"""
    try:
        text = extract_text_from_pdf(budget_path, max_pages=15)
        
        # Find GBP amounts
        gbp_matches = re.findall(r'£\s*([\d,]+(?:\.\d{2})?)', text)
        
        amounts = []
        for match in gbp_matches:
            try:
                amount = float(match.replace(',', ''))
                if 1000 < amount < 10000000:
                    amounts.append(amount)
            except:
                pass
        
        if amounts:
            return {
                'total_gbp': round(max(amounts), 2),
                'source': 'pdf',
                'amounts_found': len(amounts)
            }
        
        return {'total_gbp': None, 'note': 'No GBP in PDF'}
        
    except Exception as e:
        return {'total_gbp': None, 'error': f'pdf: {str(e)[:100]}'}

def extract_from_schedule(schedule_path: Path) -> Dict[str, Any]:
    """Extract schedule data"""
    try:
        ext = schedule_path.suffix.lower()
        
        if ext in ['.xlsx', '.xls']:
            if ext == '.xlsx' and HAS_OPENPYXL:
                wb = openpyxl.load_workbook(schedule_path, read_only=True, data_only=True)
                sheet = wb.active
                text_parts = []
                for row in sheet.iter_rows(max_row=100, values_only=True):
                    if row:
                        text_parts.append(' '.join([str(c) for c in row if c]))
                text = '\n'.join(text_parts)
                wb.close()
            elif ext == '.xls' and HAS_XLRD:
                wb = xlrd.open_workbook(schedule_path)
                sheet = wb.sheet_by_index(0)
                text_parts = []
                for row_idx in range(min(100, sheet.nrows)):
                    row = [str(sheet.cell(row_idx, col_idx).value) for col_idx in range(sheet.ncols)]
                    text_parts.append(' '.join(row))
                text = '\n'.join(text_parts)
            else:
                return {'shoot_days': None, 'error': f'Cannot read {ext}'}
        else:
            # PDF
            text = extract_text_from_pdf(schedule_path, max_pages=20)
        
        text_lower = text.lower()
        
        # Count days
        day_matches = re.findall(r'day\s+(\d+)', text_lower)
        shoot_day_matches = re.findall(r'shoot\s+day\s+(\d+)', text_lower)
        all_days = day_matches + shoot_day_matches
        shoot_days = max([int(d) for d in all_days]) if all_days else None
        
        # Call times
        call_times = re.findall(r'call[:\s]+(\d{1,2})[:\.](\d{2})', text_lower)
        
        return {
            'shoot_days': shoot_days,
            'call_times_found': len(call_times)
        }
        
    except Exception as e:
        return {'shoot_days': None, 'error': f'{str(e)[:100]}'}

def process_project(project: Dict[str, Any]) -> Dict[str, Any]:
    """Process a single project"""
    project_name = project.get('project_name', 'Unknown')
    print(f"\nProcessing: {project_name}")
    
    result = {
        'project_name': project_name,
        'client': project.get('client', ''),
        'complete': project.get('complete', False),
        'script_features': {},
        'budget_data': {},
        'schedule_data': {}
    }
    
    files = project.get('files', {})
    
    # Script
    if 'script' in files:
        script_file = files['script']
        if isinstance(script_file, dict) and 'path' in script_file:
            script_path = Path(script_file['path'])
            if script_path.exists():
                print(f"  📄 Script: {script_path.name}")
                text = extract_text_from_pdf(script_path, max_pages=20)
                result['script_features'] = extract_features_from_script(text)
    
    # Budget
    if 'budget' in files:
        budget_file = files['budget']
        if isinstance(budget_file, dict) and 'path' in budget_file:
            budget_path = Path(budget_file['path'])
            if budget_path.exists():
                print(f"  💰 Budget: {budget_path.name}")
                ext = budget_path.suffix.lower()
                if ext == '.xlsx':
                    result['budget_data'] = extract_from_xlsx_budget(budget_path)
                elif ext == '.xls':
                    result['budget_data'] = extract_from_xls_budget(budget_path)
                else:
                    result['budget_data'] = extract_from_pdf_budget(budget_path)
        elif isinstance(budget_file, list):
            for bf in budget_file:
                if isinstance(bf, dict) and 'path' in bf:
                    budget_path = Path(bf['path'])
                    if budget_path.exists():
                        print(f"  💰 Budget: {budget_path.name}")
                        ext = budget_path.suffix.lower()
                        if ext == '.xlsx':
                            result['budget_data'] = extract_from_xlsx_budget(budget_path)
                        elif ext == '.xls':
                            result['budget_data'] = extract_from_xls_budget(budget_path)
                        else:
                            result['budget_data'] = extract_from_pdf_budget(budget_path)
                        break
    
    # Schedule
    if 'schedule' in files:
        schedule_file = files['schedule']
        if isinstance(schedule_file, dict) and 'path' in schedule_file:
            schedule_path = Path(schedule_file['path'])
            if schedule_path.exists():
                print(f"  📅 Schedule: {schedule_path.name}")
                result['schedule_data'] = extract_from_schedule(schedule_path)
        elif isinstance(schedule_file, list):
            for sf in schedule_file:
                if isinstance(sf, dict) and 'path' in sf and sf.get('exists', False):
                    schedule_path = Path(sf['path'])
                    if schedule_path.exists():
                        print(f"  📅 Schedule: {schedule_path.name}")
                        result['schedule_data'] = extract_from_schedule(schedule_path)
                        break
    
    print(f"  ✅ Done")
    return result

def main():
    base_path = Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Henry-ClientDocs/reference-data"
    manifest_path = base_path / "training_data_extract.json"
    output_dir = Path.home() / "clawd/projects/Production Script Platform/production-feasibility-engine/training-data"
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"📂 Loading manifest...")
    with open(manifest_path) as f:
        manifest = json.load(f)
    
    projects = manifest.get('projects', [])
    print(f"📊 Processing {len(projects)} projects")
    print("="*60)
    
    results = []
    
    for i, project in enumerate(projects, 1):
        print(f"\n[{i}/{len(projects)}]", end=' ')
        try:
            result = process_project(project)
            results.append(result)
            
            if i % 5 == 0:
                output_path = output_dir / "training_data_partial.json"
                with open(output_path, 'w') as f:
                    json.dump(results, f, indent=2)
                print(f"  💾 Checkpoint saved")
                
        except Exception as e:
            print(f"  ❌ Failed: {e}")
            results.append({
                'project_name': project.get('project_name', 'Unknown'),
                'error': str(e)
            })
    
    print(f"\n\n{'='*60}")
    print("✅ EXTRACTION COMPLETE")
    print(f"{'='*60}")
    
    output_path = output_dir / "training_data_complete.json"
    with open(output_path, 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"\n💾 Saved: {output_path}")
    
    # Summary
    successful = sum(1 for r in results if 'error' not in r)
    with_script = sum(1 for r in results if r.get('script_features', {}).get('text_length', 0) > 100)
    with_budget = sum(1 for r in results if r.get('budget_data', {}).get('total_gbp'))
    with_schedule = sum(1 for r in results if r.get('schedule_data', {}).get('shoot_days'))
    
    print(f"\n📊 Summary:")
    print(f"  Total: {len(results)}")
    print(f"  Successful: {successful}")
    print(f"  Scripts: {with_script}")
    print(f"  Budgets: {with_budget}")
    print(f"  Schedules: {with_schedule}")
    
    # Sample budgets
    budgets_found = [r for r in results if r.get('budget_data', {}).get('total_gbp')]
    if budgets_found:
        print(f"\n💰 Budgets extracted:")
        for r in budgets_found[:10]:
            total = r['budget_data']['total_gbp']
            print(f"  {r['project_name']}: £{total:,.0f}")

if __name__ == '__main__':
    main()
