#!/usr/bin/env python3
"""
Production data extractor v3 - Fixed Excel .xls handling
Uses pdfplumber, openpyxl, and xlrd for legacy Excel files
"""

import json
import os
import sys
from pathlib import Path
from typing import Dict, Any, Optional
import re
import warnings
warnings.filterwarnings('ignore')

try:
    import pdfplumber
    import openpyxl
    import xlrd  # For legacy .xls files
    LIBRARIES_OK = True
except ImportError as e:
    print(f"❌ Missing libraries: {e}")
    print("Run: pip3 install pdfplumber openpyxl xlrd")
    LIBRARIES_OK = False
    sys.exit(1)

def extract_text_from_pdf(pdf_path: Path, max_pages: int = 10, max_chars: int = 50000) -> str:
    """Extract text from PDF using pdfplumber"""
    try:
        text_parts = []
        with pdfplumber.open(pdf_path) as pdf:
            pages_to_read = min(len(pdf.pages), max_pages)
            
            for i, page in enumerate(pdf.pages[:pages_to_read]):
                if len(''.join(text_parts)) > max_chars:
                    break
                page_text = page.extract_text()
                if page_text:
                    text_parts.append(page_text)
        
        full_text = '\n'.join(text_parts)
        return full_text[:max_chars]
        
    except Exception as e:
        return f"[PDF extraction failed: {e}]"

def extract_features_from_script(text: str) -> Dict[str, Any]:
    """Extract key features from script text"""
    text_lower = text.lower()
    
    techniques = []
    technique_keywords = {
        'moco': ['moco', 'motion control', 'mo-co'],
        'drone': ['drone', 'aerial', 'uav'],
        'tracking': ['tracking shot', 'tracking', 'dolly track'],
        'vfx': ['vfx', 'visual effects', 'green screen', 'greenscreen', 'cgi'],
        'night_shoot': ['night shoot', 'night exterior', 'night int', 'night ext'],
        'underwater': ['underwater', 'submerged'],
        'crane': ['crane shot', 'crane', 'jib'],
        'steadicam': ['steadicam', 'steadi'],
        'handheld': ['handheld', 'hand held'],
        'slow_motion': ['slow motion', 'slow-motion', 'high speed', 'phantom'],
        'time_lapse': ['time lapse', 'time-lapse', 'timelapse']
    }
    
    for tech, keywords in technique_keywords.items():
        if any(kw in text_lower for kw in keywords):
            techniques.append(tech)
    
    locations = []
    if 'studio' in text_lower or 'sound stage' in text_lower:
        locations.append('studio')
    if any(word in text_lower for word in ['exterior', 'ext.', 'outdoor', 'location']):
        locations.append('outdoor')
    if 'interior' in text_lower or 'int.' in text_lower:
        locations.append('indoor')
    
    shot_patterns = [
        r'shot\s+\d+',
        r'scene\s+\d+',
        r'sc\.\s*\d+',
        r'\d+\s*\.\s*(?:int|ext)'
    ]
    
    all_shots = []
    for pattern in shot_patterns:
        matches = re.findall(pattern, text_lower)
        all_shots.extend(matches)
    
    estimated_shots = len(set(all_shots)) if all_shots else None
    
    has_children = any(word in text_lower for word in ['child', 'kid', 'baby', 'infant'])
    has_animals = any(word in text_lower for word in ['dog', 'cat', 'horse', 'animal'])
    has_vehicles = any(word in text_lower for word in ['car', 'vehicle', 'truck', 'motorcycle'])
    
    return {
        'techniques': techniques,
        'locations': locations,
        'estimated_shots': estimated_shots,
        'has_children': has_children,
        'has_animals': has_animals,
        'has_vehicles': has_vehicles,
        'text_length': len(text)
    }

def extract_amounts_from_cells(sheet, max_rows=200) -> list:
    """Extract monetary amounts from spreadsheet cells"""
    amounts = []
    
    for row_idx in range(min(sheet.nrows, max_rows)):
        for col_idx in range(sheet.ncols):
            try:
                cell = sheet.cell(row_idx, col_idx)
                cell_value = cell.value
                
                if cell_value is None:
                    continue
                
                # Check if numeric
                if isinstance(cell_value, (int, float)):
                    if 10000 < cell_value < 10000000:
                        amounts.append(float(cell_value))
                
                # Check if string contains GBP
                elif isinstance(cell_value, str):
                    # Look for £ amounts
                    gbp_matches = re.findall(r'£\s*([\d,]+(?:\.\d{2})?)', cell_value)
                    for match in gbp_matches:
                        try:
                            amount = float(match.replace(',', ''))
                            if 1000 < amount < 10000000:
                                amounts.append(amount)
                        except:
                            pass
            except:
                continue
    
    return amounts

def extract_from_xls_budget(budget_path: Path) -> Dict[str, Any]:
    """Extract budget data from legacy .xls file using xlrd"""
    try:
        print(f"    Reading .xls file: {budget_path.name}...")
        workbook = xlrd.open_workbook(budget_path, formatting_info=False)
        sheet = workbook.sheet_by_index(0)
        
        amounts = extract_amounts_from_cells(sheet, max_rows=200)
        
        if amounts:
            total = max(amounts)
            return {
                'total_gbp': round(total, 2),
                'source': 'xls_extracted',
                'amounts_found': len(amounts)
            }
        
        return {'total_gbp': None, 'note': 'No amounts found in .xls file'}
        
    except Exception as e:
        return {'total_gbp': None, 'error': f'xls: {str(e)}'}

def extract_from_xlsx_budget(budget_path: Path) -> Dict[str, Any]:
    """Extract budget data from modern .xlsx file"""
    try:
        print(f"    Reading .xlsx file: {budget_path.name}...")
        wb = openpyxl.load_workbook(budget_path, read_only=True, data_only=True)
        sheet = wb.active
        
        amounts = []
        
        for row in sheet.iter_rows(max_row=200, values_only=True):
            if not row:
                continue
            
            for cell in row:
                if cell is None:
                    continue
                    
                # Numeric cell
                if isinstance(cell, (int, float)) and 10000 < cell < 10000000:
                    amounts.append(cell)
                
                # String cell with £
                elif isinstance(cell, str):
                    gbp_matches = re.findall(r'£\s*([\d,]+(?:\.\d{2})?)', cell)
                    for match in gbp_matches:
                        try:
                            amount = float(match.replace(',', ''))
                            if 1000 < amount < 10000000:
                                amounts.append(amount)
                        except:
                            pass
        
        wb.close()
        
        if amounts:
            total = max(amounts)
            return {
                'total_gbp': round(total, 2),
                'source': 'xlsx_extracted',
                'amounts_found': len(amounts)
            }
        
        return {'total_gbp': None, 'note': 'No amounts found in .xlsx file'}
        
    except Exception as e:
        return {'total_gbp': None, 'error': f'xlsx: {str(e)}'}

def extract_from_pdf_budget(budget_path: Path) -> Dict[str, Any]:
    """Extract budget data from PDF"""
    try:
        text = extract_text_from_pdf(budget_path, max_pages=15)
        
        gbp_matches = re.findall(r'£\s*([\d,]+(?:\.\d{2})?)', text)
        
        amounts = []
        for match in gbp_matches:
            try:
                amount = float(match.replace(',', ''))
                if 1000 < amount < 10000000:
                    amounts.append(amount)
            except:
                pass
        
        if amounts:
            total = max(amounts)
            return {
                'total_gbp': round(total, 2),
                'source': 'pdf_extracted',
                'amounts_found': len(amounts)
            }
        
        return {'total_gbp': None, 'note': 'No GBP amounts found in PDF'}
        
    except Exception as e:
        return {'total_gbp': None, 'error': f'PDF parse error: {str(e)}'}

def extract_from_schedule(schedule_path: Path) -> Dict[str, Any]:
    """Extract schedule data from PDF or Excel"""
    try:
        if schedule_path.suffix.lower() == '.xlsx':
            print(f"    Reading .xlsx schedule: {schedule_path.name}...")
            wb = openpyxl.load_workbook(schedule_path, read_only=True, data_only=True)
            sheet = wb.active
            
            text_parts = []
            for row in sheet.iter_rows(max_row=100, values_only=True):
                if row:
                    text_parts.append(' '.join([str(c) for c in row if c]))
            
            text = '\n'.join(text_parts)
            wb.close()
            
        elif schedule_path.suffix.lower() == '.xls':
            print(f"    Reading .xls schedule: {schedule_path.name}...")
            workbook = xlrd.open_workbook(schedule_path)
            sheet = workbook.sheet_by_index(0)
            
            text_parts = []
            for row_idx in range(min(100, sheet.nrows)):
                row_values = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.value:
                        row_values.append(str(cell.value))
                if row_values:
                    text_parts.append(' '.join(row_values))
            
            text = '\n'.join(text_parts)
        else:
            # PDF schedule
            text = extract_text_from_pdf(schedule_path, max_pages=20)
        
        text_lower = text.lower()
        
        # Count day references
        day_matches = re.findall(r'day\s+(\d+)', text_lower)
        shoot_day_matches = re.findall(r'shoot\s+day\s+(\d+)', text_lower)
        
        all_days = day_matches + shoot_day_matches
        shoot_days = max([int(d) for d in all_days]) if all_days else None
        
        # Look for call times
        call_time_matches = re.findall(r'call[:\s]+(\d{1,2})[:\.](\d{2})', text_lower)
        
        # Count setups
        setup_mentions = len(re.findall(r'setup|set[\s-]up', text_lower))
        
        return {
            'shoot_days': shoot_days,
            'call_times_found': len(call_time_matches),
            'setup_mentions': setup_mentions,
            'text_length': len(text)
        }
        
    except Exception as e:
        return {'shoot_days': None, 'error': f'Schedule parse error: {str(e)}'}

def process_project(project: Dict[str, Any], base_path: Path) -> Dict[str, Any]:
    """Process a single project"""
    project_name = project.get('project_name', 'Unknown')
    print(f"\n{'='*60}")
    print(f"📦 Processing: {project_name}")
    print(f"{'='*60}")
    
    result = {
        'project_name': project_name,
        'client': project.get('client', ''),
        'complete': project.get('complete', False),
        'script_features': {},
        'budget_data': {},
        'schedule_data': {}
    }
    
    files = project.get('files', {})
    
    # Process script
    if 'script' in files:
        script_file = files['script']
        if isinstance(script_file, dict) and 'path' in script_file:
            script_path = Path(script_file['path'])
            if script_path.exists():
                print(f"  📄 Script: {script_path.name}")
                text = extract_text_from_pdf(script_path, max_pages=20)
                result['script_features'] = extract_features_from_script(text)
                if result['script_features']['text_length'] > 0:
                    print(f"     ✓ Extracted {result['script_features']['text_length']} chars")
                    if result['script_features']['techniques']:
                        print(f"     ✓ Techniques: {', '.join(result['script_features']['techniques'])}")
    
    # Process budget
    if 'budget' in files:
        budget_file = files['budget']
        if isinstance(budget_file, dict) and 'path' in budget_file:
            budget_path = Path(budget_file['path'])
            if budget_path.exists():
                print(f"  💰 Budget: {budget_path.name}")
                ext = budget_path.suffix.lower()
                if ext == '.xls':
                    result['budget_data'] = extract_from_xls_budget(budget_path)
                elif ext == '.xlsx':
                    result['budget_data'] = extract_from_xlsx_budget(budget_path)
                else:
                    result['budget_data'] = extract_from_pdf_budget(budget_path)
                
                if result['budget_data'].get('total_gbp'):
                    print(f"     ✓ Total: £{result['budget_data']['total_gbp']:,.2f}")
                    
        elif isinstance(budget_file, list):
            for bf in budget_file:
                if isinstance(bf, dict) and 'path' in bf:
                    budget_path = Path(bf['path'])
                    if budget_path.exists():
                        print(f"  💰 Budget: {budget_path.name}")
                        ext = budget_path.suffix.lower()
                        if ext == '.xls':
                            result['budget_data'] = extract_from_xls_budget(budget_path)
                        elif ext == '.xlsx':
                            result['budget_data'] = extract_from_xlsx_budget(budget_path)
                        else:
                            result['budget_data'] = extract_from_pdf_budget(budget_path)
                        
                        if result['budget_data'].get('total_gbp'):
                            print(f"     ✓ Total: £{result['budget_data']['total_gbp']:,.2f}")
                            break
    
    # Process schedule
    if 'schedule' in files:
        schedule_file = files['schedule']
        if isinstance(schedule_file, dict) and 'path' in schedule_file:
            schedule_path = Path(schedule_file['path'])
            if schedule_path.exists():
                print(f"  📅 Schedule: {schedule_path.name}")
                result['schedule_data'] = extract_from_schedule(schedule_path)
                if result['schedule_data'].get('shoot_days'):
                    print(f"     ✓ Shoot days: {result['schedule_data']['shoot_days']}")
                    
        elif isinstance(schedule_file, list):
            for sf in schedule_file:
                if isinstance(sf, dict) and 'path' in sf and sf.get('exists', False):
                    schedule_path = Path(sf['path'])
                    if schedule_path.exists():
                        print(f"  📅 Schedule: {schedule_path.name}")
                        result['schedule_data'] = extract_from_schedule(schedule_path)
                        if result['schedule_data'].get('shoot_days'):
                            print(f"     ✓ Shoot days: {result['schedule_data']['shoot_days']}")
                            break
    
    print(f"  ✅ Done")
    return result

def main():
    if not LIBRARIES_OK:
        return
    
    # Paths
    base_path = Path.home() / "Library/Mobile Documents/com~apple~CloudDocs/Henry-ClientDocs/reference-data"
    manifest_path = base_path / "training_data_extract.json"
    output_dir = Path.home() / "clawd/projects/Production Script Platform/production-feasibility-engine/training-data"
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"📂 Loading manifest: {manifest_path}")
    with open(manifest_path) as f:
        manifest = json.load(f)
    
    projects = manifest.get('projects', [])
    print(f"📊 Found {len(projects)} projects to process\n")
    
    results = []
    errors = []
    
    for i, project in enumerate(projects, 1):
        print(f"\n[{i}/{len(projects)}]")
        try:
            result = process_project(project, base_path)
            results.append(result)
            
            # Save checkpoint every 5 projects
            if i % 5 == 0:
                checkpoint_path = output_dir / "training_data_checkpoint.json"
                with open(checkpoint_path, 'w') as f:
                    json.dump(results, f, indent=2)
                print(f"\n  💾 Checkpoint saved")
                
        except Exception as e:
            error_msg = f"Project {i} ({project.get('project_name', 'Unknown')}): {str(e)}"
            print(f"\n  ❌ Failed: {e}")
            errors.append(error_msg)
            results.append({
                'project_name': project.get('project_name', 'Unknown'),
                'error': str(e)
            })
    
    # Final save
    print(f"\n\n{'='*60}")
    print("✅ EXTRACTION COMPLETE")
    print(f"{'='*60}")
    
    output_path = output_dir / "training_data_complete.json"
    with open(output_path, 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"\n💾 Saved: {output_path}")
    
    # Stats
    successful = sum(1 for r in results if 'error' not in r)
    with_script = sum(1 for r in results if r.get('script_features', {}).get('text_length', 0) > 0)
    with_budget = sum(1 for r in results if r.get('budget_data', {}).get('total_gbp'))
    with_schedule = sum(1 for r in results if r.get('schedule_data', {}).get('shoot_days'))
    
    print(f"\n📊 Summary:")
    print(f"  Total: {len(results)}")
    print(f"  Successful: {successful}")
    print(f"  Scripts: {with_script}")
    print(f"  Budgets: {with_budget}")
    print(f"  Schedules: {with_schedule}")
    
    if errors:
        print(f"\n❌ Errors ({len(errors)}):")
        for err in errors[:5]:
            print(f"  - {err}")
    
    # Sample budgets
    budgets_found = [r for r in results if r.get('budget_data', {}).get('total_gbp')]
    if budgets_found:
        print(f"\n💰 Sample budgets:")
        for r in budgets_found[:10]:
            total = r['budget_data']['total_gbp']
            print(f"  - {r['project_name']}: £{total:,.2f}")
    
    # Sample schedules
    schedules_found = [r for r in results if r.get('schedule_data', {}).get('shoot_days')]
    if schedules_found:
        print(f"\n📅 Sample schedules:")
        for r in schedules_found[:10]:
            days = r['schedule_data']['shoot_days']
            print(f"  - {r['project_name']}: {days} days")

if __name__ == '__main__':
    main()
